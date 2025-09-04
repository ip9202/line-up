from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.utils.database import get_db
from app.models.lineup import Lineup
from app.models.lineup_player import LineupPlayer
from app.models.player import Player
from app.models.game import Game
from app.models.team import Team
from app.models.venue import Venue
from app.models.user import User
from app.dependencies.auth import get_current_active_user
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import traceback

router = APIRouter()

@router.get("/lineup/{lineup_id}/excel")
async def generate_lineup_excel(
    lineup_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """라인업 엑셀 생성"""
    try:
        # 라인업 정보 조회
        lineup = db.query(Lineup).filter(Lineup.id == lineup_id).first()
        if not lineup:
            raise HTTPException(status_code=404, detail="Lineup not found")
        
        # 경기 정보 조회
        game = db.query(Game).filter(Game.id == lineup.game_id).first()
        if not game:
            raise HTTPException(status_code=404, detail="Game not found")
        
        # 상대팀 정보 조회
        opponent_team = None
        if game.opponent_team_id:
            opponent_team = db.query(Team).filter(Team.id == game.opponent_team_id).first()
        
        # 경기장 정보 조회
        venue = db.query(Venue).filter(Venue.id == game.venue_id).first()
        
        # 감독 정보 조회 (선수 중에서 COACH 역할인 사람 우선)
        coach_player = db.query(Player).filter(Player.role == 'COACH').first()
        if coach_player:
            coach_name = coach_player.name
        else:
            # 선수 중에 감독이 없으면 사용자 테이블에서 찾기
            coach = db.query(User).filter(User.role == 'coach').first()
            coach_name = coach.username if coach else '감독'
        
        # 우리팀 정보 조회
        our_team = db.query(Team).filter(Team.is_active == True).first()
        team_name = our_team.name if our_team else '씨밀레'
        
        # 라인업 플레이어 정보 조회
        lineup_players = db.query(LineupPlayer).filter(
            LineupPlayer.lineup_id == lineup_id
        ).order_by(LineupPlayer.batting_order).all()
        
        # 선수 정보 조회
        player_ids = [lp.player_id for lp in lineup_players]
        players = db.query(Player).filter(Player.id.in_(player_ids)).all()
        player_dict = {p.id: p for p in players}
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "라인업"
        
        # A4 사이즈 설정 (210mm x 297mm)
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # 가로 방향
        
        # 스타일 정의
        header_font = Font(name='맑은 고딕', size=12, bold=True)
        data_font = Font(name='맑은 고딕', size=11)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # 테두리 스타일
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 헤더 배경색
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # 1. 게임 정보 테이블 (A1:B5)
        game_info_data = [
            ["팀명", team_name],
            ["감독", coach_name],
            ["날짜", f"{game.game_date.strftime('%m.%d(%a)')} {game.game_date.strftime('%H:%M')}"],
            ["구장", venue.name if venue else '미정'],
            ["상대팀", opponent_team.name if opponent_team else '미정']
        ]
        
        for row_idx, (label, value) in enumerate(game_info_data, 1):
            ws[f'A{row_idx}'] = label
            ws[f'B{row_idx}'] = value
            
            # 스타일 적용
            ws[f'A{row_idx}'].font = data_font
            ws[f'B{row_idx}'].font = data_font
            ws[f'A{row_idx}'].alignment = left_alignment
            ws[f'B{row_idx}'].alignment = left_alignment
            ws[f'A{row_idx}'].border = thin_border
            ws[f'B{row_idx}'].border = thin_border
        
        # 열 너비 설정
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        
        # 2. 라인업 테이블 (A7:F17)
        lineup_headers = ["타순", "위치", "성명", "배번", "비고"]
        for col_idx, header in enumerate(lineup_headers, 1):
            cell = ws.cell(row=7, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = header_fill
        
        # 라인업 데이터
        lineup_data = []
        
        # 타순 1-9번
        for i in range(1, 10):
            row = [str(i), "", "", "", ""]  # 타순, 위치, 성명, 배번, 비고
            for lp in lineup_players:
                if lp.batting_order == i:
                    player = player_dict.get(lp.player_id)
                    if player:
                        position = lp.position or ""
                        if position == "1B":
                            position = "1루수"
                        elif position == "2B":
                            position = "2루수"
                        elif position == "3B":
                            position = "3루수"
                        elif position == "SS":
                            position = "유격수"
                        elif position == "LF":
                            position = "좌익수"
                        elif position == "CF":
                            position = "중견수"
                        elif position == "RF":
                            position = "우익수"
                        elif position == "C":
                            position = "포수"
                        elif position == "DH":
                            position = "지명타자"
                        elif position == "P":
                            position = "투수"
                        
                        row[1] = position
                        row[2] = player.name
                        row[3] = str(player.number)
                    break
            lineup_data.append(row)
        
        # 투수 추가
        pitcher_row = ["P", "투수", "", "", ""]
        for lp in lineup_players:
            if lp.batting_order == 0:  # 투수는 0번
                player = player_dict.get(lp.player_id)
                if player:
                    pitcher_row[2] = player.name
                    pitcher_row[3] = str(player.number)
                break
        lineup_data.append(pitcher_row)
        
        # 라인업 데이터 입력
        for row_idx, row_data in enumerate(lineup_data, 8):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = center_alignment
                cell.border = thin_border
        
        # 라인업 테이블 열 너비 설정
        lineup_widths = [8, 12, 15, 8, 8]
        for col_idx, width in enumerate(lineup_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # 3. 선수 명단 테이블 (H1:K25)
        player_headers = ["번호", "성명", "배번", "비고"]
        for col_idx, header in enumerate(player_headers, 8):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = header_fill
        
        # 선수 명단 데이터
        all_players = db.query(Player).filter(Player.is_active == True).order_by(Player.number).all()
        for row_idx, player in enumerate(all_players, 2):
            player_data = [str(player.number), player.name, str(player.number), ""]
            for col_idx, value in enumerate(player_data, 8):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = center_alignment
                cell.border = thin_border
        
        # 선수 명단 테이블 열 너비 설정
        player_widths = [8, 15, 8, 8]
        for col_idx, width in enumerate(player_widths, 8):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # 엑셀 파일 저장
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            BytesIO(buffer.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=lineup_{lineup_id}.xlsx"}
        )
        
    except Exception as e:
        print(f"엑셀 생성 에러: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"엑셀 생성 실패: {str(e)}")
